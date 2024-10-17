import tkinter as tk
from tkinter import filedialog, messagebox
import pandas as pd
import os
import platform
import threading
from tkinter import font
from openpyxl.styles import Alignment, Font
from openpyxl.styles import Border, Side
from openpyxl.styles import PatternFill
from openpyxl.worksheet.page import PrintPageSetup
from datetime import datetime

# Global variable to store file path and report type
selected_file_path = None
report_type = "Relatório Completo"  # Default to Relatório Completo
split_by_name = False  # Variable for splitting reports by name
target_minutes = 510  # Default target work time in minutes
minutes_limit = 15  # Default tolerance limit in minutes
save_location = ""  # Store the selected file location


def set_report_type(selected_type):
    """
    Sets the report type based on the user's selection in the dropdown menu.
    """
    global report_type
    report_type = selected_type


def select_file():
    global selected_file_path

    # Open file dialog to select a CSV file
    file_path = filedialog.askopenfilename(
        title="Escolha o Ficheiro Exportado",
        filetypes=[("CSV Files", "*.csv")]
    )

    if file_path:
        # Display the selected file path in the label
        selected_file_path = file_path
        file_label.config(text=f"{file_path}")


def select_save_location():
    """
    Prompts the user to select the location to save the report, and
    sets the save location in the GUI.
    """
    global save_location

    # Open a file dialog to choose save location
    save_location = filedialog.asksaveasfilename(
        title="Escolha o local para guardar o relatório",
        defaultextension=".xlsx",  # Set default extension as Excel
        filetypes=[("Excel Files", "*.xlsx"), ("PDF Files", "*.pdf")]
    )

    if save_location:
        filename_entry.delete(0, tk.END)
        filename_entry.insert(0, os.path.basename(save_location))


def toggle_split_by_name():
    global split_by_name
    split_by_name = not split_by_name


def start_report_generation():
    """
    Starts the report generation in a separate thread.
    """
    # Show the loading message
    loading_label.config(text="A gerar o relatório... Por favor aguarde.")
    loading_label.update()

    # Start a new thread for report generation
    report_thread = threading.Thread(target=generate_report_button)
    report_thread.start()


def generate_report_button():
    global selected_file_path, split_by_name, report_type, target_minutes, minutes_limit, save_location

    # Get user input for target minutes and tolerance limit
    try:
        target_minutes = int(target_minutes_entry.get())
        minutes_limit = int(minutes_limit_entry.get())
    except ValueError:
        messagebox.showerror("Erro", "Por favor introduza números inteiros.")
        return

    if selected_file_path and save_location:
        try:
            # Read the selected CSV file
            df = pd.read_csv(selected_file_path)

            # Generate the report based on the user's settings
            if split_by_name:
                unique_names = df['Name'].unique()
                for name in unique_names:
                    df_name = df[df['Name'] == name]
                    report_file_path = save_location.replace('.xlsx', f'_{name}_report.xlsx')
                    generate_report(df_name, report_file_path)
                messagebox.showinfo("Sucesso", "Relatórios gerados para cada colaborador.")
            else:
                # Use the user-defined save location for the report
                report_file_path = save_location
                generate_report(df, report_file_path)
                messagebox.showinfo("Sucesso", f"Relatório gerado com sucesso: {report_file_path}")

            # Open the generated report automatically if not splitting
            if not split_by_name:
                open_report(report_file_path)

        except Exception as e:

            messagebox.showerror("Erro", f"Falha ao processar o ficheiro: {e}")
    else:
        messagebox.showwarning("Nenhum ficheiro selecionado", "Por favor selecione primeiro o ficheiro e o local para guardar.")

    # Hide the loading message after completion
    loading_label.config(text="")


def generate_report(df, output_file):
    """
    Function to process the CSV data and generate the desired Excel report
    based on the selected report type (Relatório Completo or Relatório Simples).
    """
    global report_type, target_minutes, minutes_limit

    df['Time'] = pd.to_datetime(df['Time'], errors='coerce')
    df['Date'] = df['Time'].dt.date

    simplified_columns = ['Data', 'Nome', 'Entrada', 'Saída Almoço', 'Entrada Almoço', 'Saída']
    full_columns = ['Data', 'Nome', 'Entrada', 'Saída Alm.', 'Entrada Alm.', 'Saída', 'Trabalho (min)', 'Extra/Falta']

    report_columns = full_columns if report_type == "Relatório Completo" else simplified_columns

    report_data = []
    lower_limit = target_minutes - minutes_limit
    upper_limit = target_minutes + minutes_limit

    grouped = df.groupby(['Person ID', 'Name', 'Date'])

    for (person_id, name, date), group in grouped:
        check_in = group.loc[group['Attendance Status'] == 'Check in', 'Time'].min()
        check_out = group.loc[group['Attendance Status'] == 'Check out', 'Time'].max()

        lunch_out = group.loc[group['Attendance Status'] == 'Coffee out', 'Time'].min() if 'Coffee out' in group['Attendance Status'].values else None
        lunch_in = group.loc[group['Attendance Status'] == 'Coffee in', 'Time'].max() if 'Coffee in' in group['Attendance Status'].values else None

        if pd.notnull(check_in) and pd.notnull(check_out):
            total_work_time = (check_out - check_in).total_seconds() / 60
            lunch_break_time = (lunch_in - lunch_out).total_seconds() / 60 if lunch_out and lunch_in else 0

            worked_minutes = total_work_time - lunch_break_time
            overtime_or_undertime = worked_minutes - target_minutes

            worked_minutes = round(worked_minutes, 0) if worked_minutes != 0 else ""
            overtime_or_undertime = round(overtime_or_undertime, 0) if overtime_or_undertime != 0 else ""
        else:
            worked_minutes = ""
            overtime_or_undertime = ""

        if report_type == "Relatório Completo":
            report_data.append([date, name, check_in.time() if pd.notnull(check_in) else None,
                                lunch_out.time() if pd.notnull(lunch_out) else None,
                                lunch_in.time() if pd.notnull(lunch_in) else None,
                                check_out.time() if pd.notnull(check_out) else None,
                                worked_minutes, overtime_or_undertime])
        else:
            report_data.append([date, name, check_in.time() if pd.notnull(check_in) else None,
                                lunch_out.time() if pd.notnull(lunch_out) else None,
                                lunch_in.time() if pd.notnull(lunch_in) else None,
                                check_out.time() if pd.notnull(check_out) else None])

    report_df = pd.DataFrame(report_data, columns=report_columns)

    # Save the report DataFrame to an Excel file
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        # Write the data starting from row 4
        report_df.to_excel(writer, index=False, startrow=3, sheet_name='Relatório Ponto')

        # Access the openpyxl workbook and worksheet
        worksheet = writer.sheets['Relatório Ponto']

        # Define a thin border style for all cells
        thin_border = Border(
            left=Side(border_style="thin"),
            right=Side(border_style="thin"),
            top=Side(border_style="thin"),
            bottom=Side(border_style="thin")
        )

        # Define fill styles for conditional formatting
        green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Green fill for within limit
        red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # Red fill for less than lower limit
        yellow_fill = PatternFill(start_color="FFFACD", end_color="FFFACD", fill_type="solid")  # Yellow fill for more than upper limit

        # Apply borders and conditional formatting to the overtime/under-time column
        for row in worksheet.iter_rows(min_row=4, max_row=worksheet.max_row, min_col=8, max_col=8):  # Column H (Overtime/Under-time)
            for cell in row:
                # Apply borders to all cells
                cell.border = thin_border

                # Conditionally format based on overtime/under-time logic, skipping blank cells
                if isinstance(cell.value, (int, float)):
                    if lower_limit <= target_minutes + cell.value <= upper_limit:
                        cell.fill = green_fill  # Worked within the tolerance limit
                    elif cell.value < lower_limit - target_minutes:
                        cell.fill = red_fill  # Worked less than lower limit
                    elif cell.value > upper_limit - target_minutes:
                        cell.fill = yellow_fill

        # Apply borders to all data cells and center align
        for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row, min_col=1, max_col=len(report_columns)):
            for cell in row:
                cell.font = Font(size=8)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center', vertical='center')

        # Make all cells in row 4 bold
        for cell in worksheet[4]:  # Row 4 (index 4 refers to the 4th row)
            cell.font = Font(size=8, bold=True)  # Apply bold font to all cells in row 4

        # Remove borders from any empty columns beyond the relevant columns in the Relatório Simples
        if report_type == "Relatório Simples":
            # We need to clear borders in column H and beyond (after the last relevant column in the Relatório Simples)
            for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row, min_col=len(simplified_columns) + 1, max_col=len(full_columns)):
                for cell in row:
                    cell.border = None  # Remove borders from empty cells

        # Adjust column widths based on content
        for col in worksheet.columns:
            max_length = 0
            column = col[0].column_letter  # Get the column name (A, B, C, etc.)
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            worksheet.column_dimensions[column].width = max_length

        # Set the report title in row 1
        title = "Relatório de Ponto"
        title_cell = worksheet.cell(row=1, column=1)
        title_cell.value = title
        title_cell.font = Font(size=11, bold=True)
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(report_columns))

        # Add the report metadata in row 2
        date_cell = worksheet.cell(row=2, column=1)
        date_cell.value = 'Data do Relatório'
        date_cell.font = Font(bold=True, size=8)
        worksheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=2)
        worksheet['C2'] = datetime.now().strftime('%Y-%m-%d')

        company_cell = worksheet.cell(row=2, column=5)
        company_cell.value = 'Empresa'
        company_cell.font = Font(bold=True, size=8)
        worksheet['F2'] = 'Friártico'

        # Add the shift metadata in row 3
        worksheet.merge_cells(start_row=3, start_column=1, end_row=3, end_column=2)

        first_shift_cell = worksheet.cell(row=3, column=3)
        first_shift_cell.value = 'Turno Manhã'
        first_shift_cell.font = Font(bold=True, size=8)
        worksheet.merge_cells(start_row=3, start_column=3, end_row=3, end_column=4)

        first_shift_cell = worksheet.cell(row=3, column=5)
        first_shift_cell.value = 'Turno Tarde'
        first_shift_cell.font = Font(bold=True, size=8)
        worksheet.merge_cells(start_row=3, start_column=5, end_row=3, end_column=6)

        if report_type == "Relatório Completo":
            company_cell = worksheet.cell(row=2, column=7)
            company_cell.value = 'Trabalho (min)'
            company_cell.font = Font(bold=True, size=8)
            worksheet['H2'] = target_minutes

            tolerance_cell = worksheet.cell(row=3, column=7)
            tolerance_cell.value = 'Tolerância (min)'
            tolerance_cell.font = Font(bold=True, size=8)
            worksheet['H3'] = minutes_limit

        # Set the table header row (row 4) to repeat on every printed page
        worksheet.print_title_rows = '4:4'  # This will repeat row 4 on every page when printing

        # Set up the page to print correctly
        worksheet.page_setup = PrintPageSetup()
        worksheet.page_setup.orientation = 'portrait'  # Set orientation to portrait or landscape
        worksheet.page_setup.fitToWidth = 1
        worksheet.page_setup.paperSize = 9  # Set the paper size to A4 (9 is the code for A4)


def open_report(file_path):
    """
    Function to open the generated Excel report.
    """
    try:
        # Use the appropriate command depending on the OS
        if platform.system() == "Darwin":  # macOS
            os.system(f"open '{file_path}'")
        elif platform.system() == "Windows":  # Windows
            os.startfile(file_path)
        elif platform.system() == "Linux":  # Linux
            os.system(f"xdg-open '{file_path}'")
    except Exception as e:
        messagebox.showerror("Error", f"Could not open the report: {e}")


# Set up the main window
root = tk.Tk()
root.title("Relatórios Ponto Friártico")
root.geometry("500x650")

# Create a bold font
bold_font = font.Font(family="Helvetica", size=14, weight="bold")

# Add a label to show the selected file path, with bold text and custom color
file_label = tk.Label(root, text="Nenhum ficheiro selecionado", wraplength=400, font=bold_font, fg="blue")
file_label.pack(pady=10)

# Add a button to select the CSV file
btn_select_file = tk.Button(root, text="Selecione o ficheiro exportado", command=select_file)
btn_select_file.pack(pady=10)

# Add a label for selecting the save location
filename_label = tk.Label(root, text="Nome do ficheiro e local para guardar:", font=bold_font)
filename_label.pack(pady=5)

# Add an entry widget to display the selected filename and location
filename_entry = tk.Entry(root, width=30)
filename_entry.pack(pady=5)

# Add a button to select the save location
btn_select_location = tk.Button(root, text="Escolha o local para guardar", command=select_save_location)
btn_select_location.pack(pady=5)

# Add a dropdown menu to select the report type (Relatório Completo or Relatório Simples)
report_type_label = tk.Label(root, text="Selecione o tipo de relatório:", font=bold_font)
report_type_label.pack(pady=10)

# Dropdown menu options
report_type_options = ["Relatório Completo", "Relatório Simples"]
report_type_var = tk.StringVar()
report_type_var.set(report_type_options[0])  # Default to "Relatório Completo"

# Create dropdown menu
report_type_menu = tk.OptionMenu(root, report_type_var, *report_type_options, command=lambda val: set_report_type(val))
report_type_menu.pack(pady=10)

# Add entry fields for target minutes and limit minutes
target_minutes_label = tk.Label(root, text="Tempo de trabalho (minutos):", font=bold_font)
target_minutes_label.pack(pady=10)
target_minutes_entry = tk.Entry(root)
target_minutes_entry.insert(0, str(target_minutes))
target_minutes_entry.pack(pady=5)

minutes_limit_label = tk.Label(root, text="Limite de tolerância (minutos):", font=bold_font)
minutes_limit_label.pack(pady=10)
minutes_limit_entry = tk.Entry(root)
minutes_limit_entry.insert(0, str(minutes_limit))
minutes_limit_entry.pack(pady=5)

# Add a checkbox to toggle splitting the report by name
split_by_name_var = tk.IntVar()
split_by_name_checkbox = tk.Checkbutton(root, text="Criar relatórios por nome de colaborador", variable=split_by_name_var, command=toggle_split_by_name)
split_by_name_checkbox.pack(pady=10)

# Add a label to display loading info
loading_label = tk.Label(root, text="", wraplength=400, fg="green")
loading_label.pack(pady=10)

# Add a button to generate the report
btn_generate_report = tk.Button(root, text="Gerar Relatórios", command=start_report_generation)
btn_generate_report.pack(pady=10)

# Run the application
root.mainloop()
